from __future__ import annotations

import io
import os
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import List, Optional

import pandas as pd
from flask import Flask, Response, redirect, render_template_string, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "mission-freight-invoice-tool")


CHARGE_KEYWORDS = [
    "import warehouse charges",
    "handling",
    "handling charges",
    "handling fee",
]


@dataclass
class InvoiceResult:
    bestandsnaam: str
    factuurnummer: Optional[str]
    awb_nummer: Optional[str]
    totaal_kg: Optional[float]
    charges_eur: Optional[float]
    prijs_per_kg_eur: Optional[float]
    status: str


def normalize_spaces(text: str) -> str:
    text = text.replace("\xa0", " ").replace("\u200b", " ")
    text = re.sub(r"[ \t]+", " ", text)
    return text


def parse_decimal_eu(value: str) -> Optional[Decimal]:
    cleaned = value.strip().replace("EUR", "").replace("€", "").replace(" ", "")
    if not cleaned:
        return None

    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")

    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def extract_text_from_pdf_bytes(data: bytes) -> str:
    if fitz is None:
        raise RuntimeError("PyMuPDF is niet beschikbaar. Voeg pymupdf toe aan requirements.txt.")
    with fitz.open(stream=data, filetype="pdf") as doc:
        return "\n".join(page.get_text("text") for page in doc)


def extract_words_from_pdf_bytes(data: bytes):
    if fitz is None:
        raise RuntimeError("PyMuPDF is niet beschikbaar. Voeg pymupdf toe aan requirements.txt.")
    words = []
    with fitz.open(stream=data, filetype="pdf") as doc:
        for page_index, page in enumerate(doc):
            for w in page.get_text("words"):
                words.append((page_index, *w))
    return words


def find_invoice_number(text: str) -> Optional[str]:
    patterns = [
        r"FACTUURNUMMER\s+([0-9]{5,})",
        r"FACTUURNUMMER\s*[:\-]?\s*([A-Z0-9\-]{5,})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return None


def find_awb_number(text: str) -> Optional[str]:
    patterns = [
        r"AWB\s*NUMMER\s*[:\-]?\s*([0-9]{3}-[0-9]{8,})",
        r"\b([0-9]{3}-[0-9]{8,})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return None


def find_total_weight_kg_from_words(words) -> Optional[Decimal]:
    """
    Zoekt letterlijk het woord BRUTO in de PDF-layout en pakt
    het eerste numerieke woord daaronder in ongeveer dezelfde kolom.
    """
    if not words:
        return None

    bruto_words = []
    for item in words:
        page_index, x0, y0, x1, y1, text, block_no, line_no, word_no = item
        if str(text).strip().lower() == "bruto":
            bruto_words.append(item)

    if not bruto_words:
        return None

    bruto_word = sorted(bruto_words, key=lambda w: (w[0], w[3], w[2]))[-1]
    b_page, bx0, by0, bx1, by1, _, _, _, _ = bruto_word
    bruto_center_x = (bx0 + bx1) / 2

    candidates = []
    for item in words:
        page_index, x0, y0, x1, y1, text, block_no, line_no, word_no = item

        if page_index != b_page:
            continue

        token = str(text).strip()
        if not re.fullmatch(r"\d+(?:[.,]\d+)?", token):
            continue

        center_x = (x0 + x1) / 2

        if y0 <= by1:
            continue

        if abs(center_x - bruto_center_x) > 100:
            continue

        value = parse_decimal_eu(token)
        if value is None:
            continue

        candidates.append((y0, value))

    if candidates:
        candidates.sort(key=lambda item: item[0])
        return candidates[0][1]

    return None


def sum_import_warehouse_charges(text: str) -> Optional[Decimal]:
    """
    Verbeterde versie:
    - zoekt in de hele genormaliseerde tekst
    - werkt ook als PDF de regel over meerdere stukjes splitst
    - pakt zowel import warehouse charges als handling-varianten
    """
    total = Decimal("0")
    found = False

    normalized = normalize_spaces(text).lower()
    normalized = re.sub(r"\s+", " ", normalized)

    matches = re.findall(
        r"(import warehouse charges|handling|handling charges|handling fee).*?([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{2}))",
        normalized,
        re.IGNORECASE,
    )

    for _, amount_str in matches:
        amount = parse_decimal_eu(amount_str)
        if amount is not None:
            total += amount
            found = True

    return total if found else None


def parse_invoice(file_name: str, data: bytes) -> InvoiceResult:
    try:
        text = extract_text_from_pdf_bytes(data)
        words = extract_words_from_pdf_bytes(data)

        factuurnummer = find_invoice_number(text)
        awb_nummer = find_awb_number(text)
        totaal_kg = find_total_weight_kg_from_words(words)
        charges = sum_import_warehouse_charges(text)

        prijs_per_kg = None
        if charges is not None and totaal_kg not in (None, Decimal("0")):
            prijs_per_kg = charges / totaal_kg

        missing = []
        if not factuurnummer:
            missing.append("factuurnummer")
        if not awb_nummer:
            missing.append("AWB nummer")
        if totaal_kg is None:
            missing.append("totaal kg")
        if charges is None:
            missing.append("Import warehouse charges")

        status = "OK" if not missing else f"Ontbreekt: {', '.join(missing)}"

        return InvoiceResult(
            bestandsnaam=file_name,
            factuurnummer=factuurnummer,
            awb_nummer=awb_nummer,
            totaal_kg=float(totaal_kg) if totaal_kg is not None else None,
            charges_eur=float(charges) if charges is not None else None,
            prijs_per_kg_eur=float(prijs_per_kg) if prijs_per_kg is not None else None,
            status=status,
        )

    except Exception as exc:
        return InvoiceResult(
            bestandsnaam=file_name,
            factuurnummer=None,
            awb_nummer=None,
            totaal_kg=None,
            charges_eur=None,
            prijs_per_kg_eur=None,
            status=f"Fout bij verwerken: {exc}",
        )


def dataframe_from_results(results: List[InvoiceResult]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Factuurnummer": r.factuurnummer,
                "AWB nummer": r.awb_nummer,
                "Totaal kg": r.totaal_kg,
                "Charges (EUR)": r.charges_eur,
                "Prijs per kg (EUR)": r.prijs_per_kg_eur,
                "Bestandsnaam": r.bestandsnaam,
                "Status": r.status,
            }
            for r in results
        ]
    )


def build_excel_bytes(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultaten"

    headers = [
        "Factuurnummer",
        "AWB nummer",
        "Totaal kg",
        "Charges (EUR)",
        "Prijs per kg (EUR)",
    ]
    ws.append(headers)

    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    for _, row in df.iterrows():
        ws.append(
            [
                row.get("Factuurnummer"),
                row.get("AWB nummer"),
                row.get("Totaal kg"),
                row.get("Charges (EUR)"),
                row.get("Prijs per kg (EUR)"),
            ]
        )

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = "0.00000"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def load_session_df() -> pd.DataFrame:
    raw = session.get("results_json")
    if not raw:
        return pd.DataFrame(
            columns=[
                "Factuurnummer",
                "AWB nummer",
                "Totaal kg",
                "Charges (EUR)",
                "Prijs per kg (EUR)",
                "Bestandsnaam",
                "Status",
            ]
        )
    return pd.read_json(io.StringIO(raw))


HTML = """
<!doctype html>
<html lang="nl">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
  <title>Mission Freight factuur uitlezer</title>
  <style>
    :root {
      --bg: #f5f7fb;
      --card: #ffffff;
      --text: #14213d;
      --muted: #5b6475;
      --primary: #2563eb;
      --border: #dbe2f0;
      --ok: #117a37;
      --warn: #b26a00;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif;
      background: var(--bg);
      color: var(--text);
    }
    .wrap {
      max-width: 980px;
      margin: 0 auto;
      padding: 18px;
    }
    .card {
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 18px;
      padding: 20px;
      box-shadow: 0 8px 24px rgba(20, 33, 61, 0.06);
      margin-bottom: 16px;
    }
    h1 {
      font-size: 1.6rem;
      margin: 0 0 10px;
    }
    h2 {
      font-size: 1.2rem;
      margin: 0 0 12px;
    }
    p {
      margin: 0 0 10px;
      color: var(--muted);
      line-height: 1.5;
    }
    .chips {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-top: 10px;
    }
    .chip {
      background: #eef3ff;
      color: var(--primary);
      padding: 8px 10px;
      border-radius: 999px;
      font-size: 13px;
    }
    .dropzone {
      border: 2px dashed var(--primary);
      border-radius: 18px;
      padding: 24px;
      text-align: center;
      background: #f8fbff;
    }
    .dropzone.dragover {
      background: #eef5ff;
    }
    input[type=file] {
      width: 100%;
      margin-top: 12px;
      font-size: 16px;
    }
    .btn {
      display: inline-block;
      width: 100%;
      border: 0;
      border-radius: 14px;
      padding: 14px 16px;
      font-size: 16px;
      font-weight: 600;
      text-decoration: none;
      text-align: center;
      margin-top: 12px;
      cursor: pointer;
    }
    .btn-primary {
      background: var(--primary);
      color: white;
    }
    .btn-secondary {
      background: #eef3ff;
      color: var(--primary);
    }
    .result-card {
      border: 1px solid var(--border);
      border-radius: 14px;
      padding: 14px;
      margin-bottom: 12px;
      background: #fcfdff;
    }
    .label {
      font-size: 12px;
      color: var(--muted);
      margin-bottom: 4px;
    }
    .value {
      font-size: 16px;
      font-weight: 600;
      margin-bottom: 10px;
      word-break: break-word;
    }
    .status-ok {
      color: var(--ok);
      font-weight: 700;
    }
    .status-warn {
      color: var(--warn);
      font-weight: 700;
    }
    table {
      width: 100%;
      border-collapse: collapse;
      font-size: 14px;
    }
    th, td {
      padding: 12px 10px;
      border-bottom: 1px solid var(--border);
      text-align: left;
      vertical-align: top;
    }
    th {
      color: var(--muted);
      font-weight: 600;
    }
    .desktop-only { display: none; }
    @media (min-width: 760px) {
      .btn {
        width: auto;
        min-width: 220px;
      }
      .desktop-only { display: block; }
      .mobile-only { display: none; }
    }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <h1>Mission Freight factuur uitlezer</h1>
      <p>Upload één of meer Mission Freight PDF-facturen. De tool leest automatisch het factuurnummer, het AWB nummer, het BRUTO gewicht en de charges. Daarna berekent hij de prijs per kg.</p>
      <div class="chips">
        <span class="chip">Mission Freight template</span>
        <span class="chip">Meerdere PDF's tegelijk</span>
        <span class="chip">Excel download</span>
        <span class="chip">iPhone vriendelijk</span>
      </div>
    </div>

    <div class="card">
      <form method="post" action="{{ url_for('upload') }}" enctype="multipart/form-data">
        <div id="dropzone" class="dropzone">
          <strong>Upload je PDF-facturen</strong>
          <p>Kies één of meerdere bestanden vanaf je iPhone, iPad of computer</p>
          <input id="files" type="file" name="files" accept="application/pdf" multiple required>
        </div>
        <button class="btn btn-primary" type="submit">Verwerk facturen</button>
      </form>
      {% if has_results %}
      <a class="btn btn-secondary" href="{{ url_for('download_excel') }}">Download Excel</a>
      {% endif %}
    </div>

    {% if has_results %}
    <div class="card mobile-only">
      <h2>Resultaten</h2>
      {% for row in rows %}
      <div class="result-card">
        <div class="label">Factuurnummer</div>
        <div class="value">{{ row['Factuurnummer'] or '—' }}</div>

        <div class="label">AWB nummer</div>
        <div class="value">{{ row['AWB nummer'] or '—' }}</div>

        <div class="label">Totaal kg</div>
        <div class="value">{% if row['Totaal kg'] == row['Totaal kg'] and row['Totaal kg'] is not none %}{{ '%.3f'|format(row['Totaal kg']) }}{% else %}—{% endif %}</div>

        <div class="label">Charges (EUR)</div>
        <div class="value">{% if row['Charges (EUR)'] == row['Charges (EUR)'] and row['Charges (EUR)'] is not none %}{{ '%.2f'|format(row['Charges (EUR)']) }}{% else %}—{% endif %}</div>

        <div class="label">Prijs per kg (EUR)</div>
        <div class="value">{% if row['Prijs per kg (EUR)'] == row['Prijs per kg (EUR)'] and row['Prijs per kg (EUR)'] is not none %}{{ '%.5f'|format(row['Prijs per kg (EUR)']) }}{% else %}—{% endif %}</div>

        <div class="label">Status</div>
        <div class="value {% if row['Status'] == 'OK' %}status-ok{% else %}status-warn{% endif %}">{{ row['Status'] }}</div>
      </div>
      {% endfor %}
    </div>

    <div class="card desktop-only">
      <h2>Resultaten</h2>
      <table>
        <thead>
          <tr>
            <th>Factuurnummer</th>
            <th>AWB nummer</th>
            <th>Totaal kg</th>
            <th>Charges (EUR)</th>
            <th>Prijs per kg (EUR)</th>
            <th>Status</th>
          </tr>
        </thead>
        <tbody>
          {% for row in rows %}
          <tr>
            <td>{{ row['Factuurnummer'] or '—' }}</td>
            <td>{{ row['AWB nummer'] or '—' }}</td>
            <td>{% if row['Totaal kg'] == row['Totaal kg'] and row['Totaal kg'] is not none %}{{ '%.3f'|format(row['Totaal kg']) }}{% else %}—{% endif %}</td>
            <td>{% if row['Charges (EUR)'] == row['Charges (EUR)'] and row['Charges (EUR)'] is not none %}{{ '%.2f'|format(row['Charges (EUR)']) }}{% else %}—{% endif %}</td>
            <td>{% if row['Prijs per kg (EUR)'] == row['Prijs per kg (EUR)'] and row['Prijs per kg (EUR)'] is not none %}{{ '%.5f'|format(row['Prijs per kg (EUR)']) }}{% else %}—{% endif %}</td>
            <td class="{% if row['Status'] == 'OK' %}status-ok{% else %}status-warn{% endif %}">{{ row['Status'] }}</td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
    {% endif %}
  </div>

  <script>
    const dropzone = document.getElementById('dropzone');
    const fileInput = document.getElementById('files');

    ['dragenter', 'dragover'].forEach(evt => {
      dropzone.addEventListener(evt, e => {
        e.preventDefault();
        e.stopPropagation();
        dropzone.classList.add('dragover');
      });
    });

    ['dragleave', 'drop'].forEach(evt => {
      dropzone.addEventListener(evt, e => {
        e.preventDefault();
        e.stopPropagation();
        dropzone.classList.remove('dragover');
      });
    });

    dropzone.addEventListener('drop', e => {
      if (e.dataTransfer.files.length) {
        fileInput.files = e.dataTransfer.files;
      }
    });
  </script>
</body>
</html>
"""


@app.get("/")
def index():
    df = load_session_df()
    rows = df.to_dict(orient="records") if not df.empty else []
    return render_template_string(HTML, rows=rows, has_results=bool(rows))


@app.post("/upload")
def upload():
    uploaded_files = request.files.getlist("files")
    results: List[InvoiceResult] = []

    for file in uploaded_files:
        if not file.filename.lower().endswith(".pdf"):
            continue
        results.append(parse_invoice(file.filename, file.read()))

    df = dataframe_from_results(results)
    session["results_json"] = df.to_json(orient="records")
    return redirect(url_for("index"))


@app.get("/download_excel")
def download_excel():
    df = load_session_df()
    excel_bytes = build_excel_bytes(df)
    return send_file(
        io.BytesIO(excel_bytes),
        as_attachment=True,
        download_name="mission_freight_resultaten.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/health")
def health():
    return Response("ok", mimetype="text/plain")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "10000"))
    app.run(host="0.0.0.0", port=port, debug=False)
